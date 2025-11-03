# excel_utils.py
import pandas as pd
import openpyxl
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st


def read_excel(uploaded_file):
    return pd.read_excel(uploaded_file)


@st.cache_data
def update_excel_with_results(uploaded_file, results_df):
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    start_col = ws.max_column + 2
    new_header = list(results_df.columns)

    for i, col_name in enumerate(new_header, start=start_col):
        ws.cell(row=1, column=i, value=col_name)

    for r_idx, row in enumerate(dataframe_to_rows(results_df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
